import pandas as pd
import numpy as np
import math
import os
import matplotlib.pyplot as plt
import seaborn as sns
from matplotlib.patches import Patch
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# Example file paths (update these with actual paths)
PATH_TO_TRAIN_DATA = '../data/train.txt'  # Training data file
FILE_PATH = '../data/data_expanded.xlsx'  # Excel file containing the 'Text' and 'Category' columns
PATH_TO_SAVE_MODEL = './gib_model.pki'

# Define accepted characters and position dictionary
accepted_chars = 'abcdefghijklmnopqrstuvwxyz '
pos = dict([(char, idx) for idx, char in enumerate(accepted_chars)])

def normalize(line):
    """Return only the subset of chars from accepted_chars."""
    return [c.lower() for c in line if c.lower() in accepted_chars]

def ngram(n, l):
    """Return all n-grams from l after normalizing"""
    filtered = normalize(l)
    for start in range(0, len(filtered) - n + 1):
        yield ''.join(filtered[start:start + n])

def read_file(file_path):
    """Read file with multiple encoding attempts"""
    encodings = ['utf-8', 'latin-1', 'cp1252', 'iso-8859-1']
    
    for encoding in encodings:
        try:
            with open(file_path, 'r', encoding=encoding) as file:
                return file.readlines()
        except UnicodeDecodeError:
            continue
    
    raise UnicodeDecodeError(f"Unable to read file {file_path} with any of the attempted encodings")

# Initialize counts matrix with smoothing
k = len(accepted_chars)
counts = [[10 for i in range(k)] for i in range(k)]

# Calculate occurrences for training data (train.txt)
for line in read_file(PATH_TO_TRAIN_DATA):
    for a, b in ngram(2, line):
        counts[pos[a]][pos[b]] += 1

# Calculate Log Probabilities
for i, row in enumerate(counts):
    s = float(sum(row))
    for j in range(len(row)):
        row[j] = math.log(row[j] / s)

def avg_transition_prob(l, log_prob_mat):
    """Return the average transition probability from l through log_prob_mat."""
    log_prob = 0.0
    transition_ct = 0
    for a, b in ngram(2, l):
        log_prob += log_prob_mat[pos[a]][pos[b]]
        transition_ct += 1
    return math.exp(log_prob / (transition_ct or 1))

# Load the data from the Excel file
data_df = pd.read_excel(FILE_PATH)

# Initialize lists to store probabilities
all_texts = []
all_categories = []
all_probs = []

# Calculate probabilities for each text record in the Excel file
data_df['Probability Score'] = data_df['Text'].apply(lambda x: avg_transition_prob(x, counts))

# Add all texts and categories to the lists
for idx, row in data_df.iterrows():
    all_texts.append(row['Text'])
    all_categories.append(row['Category'])
    all_probs.append(row['Probability Score'])

# Create the Data DataFrame for storing the results
results_df = pd.DataFrame({
    'Text': all_texts,
    'Category': all_categories,
    'Probability Score': all_probs
})



# Create a new workbook and set up the worksheet
wb = Workbook()
sheet = wb.active
sheet.title = "Data"

# Add headers to the Excel sheet
headers = ['Text', 'Category', 'Probability Score']
for col_idx, header in enumerate(headers, 1):
    cell = sheet.cell(row=1, column=col_idx, value=header)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center')
    cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

# Add data rows
for row_idx, (_, row) in enumerate(results_df.iterrows(), 2):
    sheet.cell(row=row_idx, column=1, value=row['Text'])
    sheet.cell(row=row_idx, column=2, value=row['Category'])
    sheet.cell(row=row_idx, column=3, value=row['Probability Score'])

# Save the workbook
excel_filename = 'outputs/markov_results.xlsx'
wb.save(excel_filename)
print(f"Data has been written to {excel_filename}")

